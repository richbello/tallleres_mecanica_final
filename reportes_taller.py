# reportes_taller.py
# Módulo de reportes ejecutivos: ventas, órdenes, cartera y compras
# Mantiene estilo: fondo oscuro (#0f172a), paneles (#1e293b), botones naranjas ("Menu.TButton")
# Gráficos con matplotlib embebidos en Tkinter

import os
import json
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl

# Gráficos
import matplotlib
matplotlib.use("Agg")  # backend no interactivo para evitar conflictos
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

BASE_DIR = r"C:\RICHARD\RB\2025\Taller_mecánica"
VENTAS_FILE = os.path.join(BASE_DIR, "ventas.json")
ORDENES_FILE = os.path.join(BASE_DIR, "ordenes_taller.json")
CARTERA_FILE = os.path.join(BASE_DIR, "cartera.json")
COMPRAS_FILE = os.path.join(BASE_DIR, "compras.json")
EXPORT_XLSX = os.path.join(BASE_DIR, "reportes_ejecutivos.xlsx")

def _ensure_base_dir():
    if not os.path.exists(BASE_DIR):
        os.makedirs(BASE_DIR, exist_ok=True)

def _load_json(path):
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def _sum(values):
    return round(sum(values), 2)

def _ventas_stats():
    ventas = _load_json(VENTAS_FILE)
    total = _sum([v.get("total", 0) for v in ventas])
    count = len(ventas)
    por_cliente = {}
    for v in ventas:
        c = v.get("cliente", "N/A")
        por_cliente[c] = por_cliente.get(c, 0) + v.get("total", 0)
    return {"total": total, "count": count, "por_cliente": por_cliente, "raw": ventas}

def _ordenes_stats():
    ordenes = _load_json(ORDENES_FILE)
    count = len(ordenes)
    por_estado = {}
    por_servicio = {}
    total_ordenes = _sum([o.get("total", 0) for o in ordenes])
    for o in ordenes:
        e = o.get("estado", "N/A")
        s = o.get("servicio", "N/A")
        por_estado[e] = por_estado.get(e, 0) + 1
        por_servicio[s] = por_servicio.get(s, 0) + 1
    return {"count": count, "por_estado": por_estado, "por_servicio": por_servicio, "total": total_ordenes, "raw": ordenes}

def _cartera_stats():
    cartera = _load_json(CARTERA_FILE)
    total_facturas = _sum([c.get("valor_factura", 0) for c in cartera])
    total_abonos = _sum([_sum([a.get("monto", 0) for a in c.get("abonos", [])]) for c in cartera])
    saldo_total = _sum([c.get("saldo", 0) for c in cartera])
    vencidas = sum(1 for c in cartera if c.get("estado") == "Vencida")
    return {"total_facturas": total_facturas, "total_abonos": total_abonos, "saldo_total": saldo_total, "vencidas": vencidas, "raw": cartera}

def _compras_stats():
    compras = _load_json(COMPRAS_FILE)
    total = _sum([c.get("total", 0) for c in compras])
    count = len(compras)
    por_proveedor = {}
    for c in compras:
        p = c.get("proveedor", "N/A")
        por_proveedor[p] = por_proveedor.get(p, 0) + c.get("total", 0)
    return {"total": total, "count": count, "por_proveedor": por_proveedor, "raw": compras}

def _exportar_excel(resumen):
    _ensure_base_dir()
    wb = openpyxl.Workbook()

    # Ventas
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Fecha exportación", datetime.now().isoformat()])
    ws.append([])
    ws.append(["Total ventas", resumen["ventas"]["total"]])
    ws.append(["Cantidad ventas", resumen["ventas"]["count"]])
    ws.append([])
    ws.append(["Cliente", "Total"])
    for c, t in resumen["ventas"]["por_cliente"].items():
        ws.append([c, t])

    # Órdenes
    ws2 = wb.create_sheet("Órdenes")
    ws2.append(["Total órdenes (COP)", resumen["ordenes"]["total"]])
    ws2.append(["Cantidad órdenes", resumen["ordenes"]["count"]])
    ws2.append([])
    ws2.append(["Estado", "Cantidad"])
    for e, q in resumen["ordenes"]["por_estado"].items():
        ws2.append([e, q])
    ws2.append([])
    ws2.append(["Servicio", "Cantidad"])
    for s, q in resumen["ordenes"]["por_servicio"].items():
        ws2.append([s, q])

    # Cartera
    ws3 = wb.create_sheet("Cartera")
    ws3.append(["Total facturas", resumen["cartera"]["total_facturas"]])
    ws3.append(["Total abonos", resumen["cartera"]["total_abonos"]])
    ws3.append(["Saldo total", resumen["cartera"]["saldo_total"]])
    ws3.append(["Cuentas vencidas", resumen["cartera"]["vencidas"]])

    # Compras
    ws4 = wb.create_sheet("Compras")
    ws4.append(["Total compras", resumen["compras"]["total"]])
    ws4.append(["Cantidad compras", resumen["compras"]["count"]])
    ws4.append([])
    ws4.append(["Proveedor", "Total"])
    for p, t in resumen["compras"]["por_proveedor"].items():
        ws4.append([p, t])

    wb.save(EXPORT_XLSX)

class ReportesTallerApp:
    def __init__(self, root):
        _ensure_base_dir()
        self.root = root
        self.root.title("📈 Reportes Ejecutivos - Taller Mecánico")
        self.root.geometry("1150x720")
        self.root.configure(bg="#0f172a")

        self._setup_styles()
        self._build_ui()
        self._load_data()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", background="#0f172a", foreground="#e2e8f0")
        style.configure("Title.TLabel", background="#0f172a", foreground="#e2e8f0", font=("Segoe UI", 16, "bold"))
        style.configure("Menu.TButton", background="#f59e0b", foreground="#111827", font=("Segoe UI Semibold", 11), padding=6)
        style.map("Menu.TButton", background=[("active", "#fbbf24")])
        style.configure("Card.TFrame", background="#1e293b")
        style.configure("Treeview", background="#0b1220", foreground="#e2e8f0", rowheight=26, fieldbackground="#0b1220")

    def _build_ui(self):
        ttk.Label(self.root, text="Reportes ejecutivos", style="Title.TLabel").pack(anchor="w", padx=12, pady=10)

        main = tk.Frame(self.root, bg="#0f172a")
        main.pack(fill="both", expand=True, padx=12, pady=12)

        # Izquierda: KPIs y acciones
        left = tk.Frame(main, bg="#1e293b")
        left.pack(side="left", fill="y", padx=10, pady=10)

        self.kpi_text = tk.Text(left, width=42, height=18, bg="#0b1220", fg="#e2e8f0")
        self.kpi_text.pack(padx=10, pady=10)

        ttk.Button(left, text="🔄 Actualizar", style="Menu.TButton", command=self._load_data).pack(padx=10, pady=6, anchor="w")
        ttk.Button(left, text="📊 Exportar Excel", style="Menu.TButton", command=self._exportar).pack(padx=10, pady=6, anchor="w")

        # Derecha: gráficos
        right = tk.Frame(main, bg="#1e293b")
        right.pack(side="right", fill="both", expand=True, padx=10, pady=10)

        # Figura matplotlib
        self.fig = Figure(figsize=(8, 5), facecolor="#0f172a")
        self.ax1 = self.fig.add_subplot(221, facecolor="#0f172a")
        self.ax2 = self.fig.add_subplot(222, facecolor="#0f172a")
        self.ax3 = self.fig.add_subplot(223, facecolor="#0f172a")
        self.ax4 = self.fig.add_subplot(224, facecolor="#0f172a")

        for ax in [self.ax1, self.ax2, self.ax3, self.ax4]:
            ax.tick_params(colors="#e2e8f0")
            ax.spines["bottom"].set_color("#94a3b8")
            ax.spines["top"].set_color("#94a3b8")
            ax.spines["left"].set_color("#94a3b8")
            ax.spines["right"].set_color("#94a3b8")

        self.canvas = FigureCanvasTkAgg(self.fig, master=right)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

    def _load_data(self):
        v = _ventas_stats()
        o = _ordenes_stats()
        c = _cartera_stats()
        p = _compras_stats()

        # KPIs
        self.kpi_text.config(state="normal")
        self.kpi_text.delete("1.0", "end")
        lines = [
            f"Ventas: ${v['total']:,}  |  Registros: {v['count']}",
            f"Órdenes: ${o['total']:,}  |  Cantidad: {o['count']}",
            f"Cartera: Facturas ${c['total_facturas']:,}  |  Abonos ${c['total_abonos']:,}  |  Saldo ${c['saldo_total']:,}  |  Vencidas: {c['vencidas']}",
            f"Compras: ${p['total']:,}  |  Cantidad: {p['count']}",
            "",
            "Top clientes (ventas):"
        ]
        top_clientes = sorted(v["por_cliente"].items(), key=lambda x: x[1], reverse=True)[:5]
        for cli, tot in top_clientes:
            lines.append(f"• {cli}: ${tot:,}")
        lines.append("")
        lines.append("Órdenes por estado:")
        for e, q in o["por_estado"].items():
            lines.append(f"• {e}: {q}")
        self.kpi_text.insert("1.0", "\n".join(lines))
        self.kpi_text.config(state="disabled")

        # Gráfico 1: Ventas por top clientes
        self.ax1.clear()
        labels = [x[0] for x in top_clientes] or ["Sin datos"]
        vals = [x[1] for x in top_clientes] or [0]
        self.ax1.bar(labels, vals, color="#f59e0b")
        self.ax1.set_title("Top clientes (ventas)", color="#e2e8f0")

        # Gráfico 2: Órdenes por estado
        self.ax2.clear()
        labels2 = list(o["por_estado"].keys()) or ["Sin datos"]
        vals2 = list(o["por_estado"].values()) or [0]
        self.ax2.bar(labels2, vals2, color="#f59e0b")
        self.ax2.set_title("Órdenes por estado", color="#e2e8f0")

        # Gráfico 3: Cartera (facturas vs abonos vs saldo)
        self.ax3.clear()
        self.ax3.bar(["Facturas", "Abonos", "Saldo"], [c["total_facturas"], c["total_abonos"], c["saldo_total"]], color="#f59e0b")
        self.ax3.set_title("Cartera (totales)", color="#e2e8f0")

        # Gráfico 4: Compras por proveedor (top 5)
        self.ax4.clear()
        top_prov = sorted(p["por_proveedor"].items(), key=lambda x: x[1], reverse=True)[:5]
        labels4 = [x[0] for x in top_prov] or ["Sin datos"]
        vals4 = [x[1] for x in top_prov] or [0]
        self.ax4.bar(labels4, vals4, color="#f59e0b")
        self.ax4.set_title("Compras por proveedor", color="#e2e8f0")

        self.fig.tight_layout()
        self.canvas.draw()

    def _exportar(self):
        resumen = {
            "ventas": _ventas_stats(),
            "ordenes": _ordenes_stats(),
            "cartera": _cartera_stats(),
            "compras": _compras_stats()
        }
        try:
            _exportar_excel(resumen)
            messagebox.showinfo("Exportado", f"Reportes exportados a:\n{EXPORT_XLSX}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {e}")

if __name__ == "__main__":
    _ensure_base_dir()
    root = tk.Tk()
    app = ReportesTallerApp(root)
    root.mainloop()
